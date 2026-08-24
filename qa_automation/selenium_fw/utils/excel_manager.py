import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from config.settings import REPORTS_DIR

def generate_excel_reports(results):
    excel_dir = os.path.join(REPORTS_DIR, "Excel")
    os.makedirs(excel_dir, exist_ok=True)
    
    # 1. Automation_Test_Report.xlsx
    wb = openpyxl.Workbook()
    
    # Sheet 1: Executed Test Cases
    ws_executed = wb.active
    ws_executed.title = "Executed Test Cases"
    headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time", "Priority", "Failure Reason"]
    ws_executed.append(headers)
    
    # Format headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws_executed.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        
    for res in results:
        ws_executed.append([
            res["test_id"], res["module"], res["test_name"], 
            res["status"], res["execution_time"], res["priority"], res["failure_reason"]
        ])
        
    # Sheet 2: Passed Tests
    ws_passed = wb.create_sheet(title="Passed Tests")
    ws_passed.append(headers)
    
    # Sheet 3: Failed Tests
    ws_failed = wb.create_sheet(title="Failed Tests")
    ws_failed.append(headers)
    
    # Sheet 4: Skipped Tests
    ws_skipped = wb.create_sheet(title="Skipped Tests")
    ws_skipped.append(headers)
    
    passed_count = 0
    failed_count = 0
    skipped_count = 0
    
    for res in results:
        row = [res["test_id"], res["module"], res["test_name"], res["status"], res["execution_time"], res["priority"], res["failure_reason"]]
        if res["status"] == "Passed":
            ws_passed.append(row)
            passed_count += 1
        elif res["status"] == "Failed":
            ws_failed.append(row)
            failed_count += 1
        else:
            ws_skipped.append(row)
            skipped_count += 1
            
    # Sheet 5: Execution Metrics
    ws_metrics = wb.create_sheet(title="Execution Metrics")
    ws_metrics.append(["Metric", "Value"])
    total = len(results)
    ws_metrics.append(["Total Tests", total])
    ws_metrics.append(["Passed", passed_count])
    ws_metrics.append(["Failed", failed_count])
    ws_metrics.append(["Skipped", skipped_count])
    pass_rate = (passed_count / total * 100) if total > 0 else 0
    ws_metrics.append(["Pass Rate (%)", f"{pass_rate:.2f}%"])
    
    # Sheet 6: Defect Summary
    ws_defect = wb.create_sheet(title="Defect Summary")
    ws_defect.append(["Test ID", "Test Name", "Module", "Failure Reason"])
    for res in results:
        if res["status"] == "Failed":
            ws_defect.append([res["test_id"], res["test_name"], res["module"], res["failure_reason"]])
    
    wb.save(os.path.join(excel_dir, "Automation_Test_Report.xlsx"))
    
    # Also save separate files as requested
    wb_failed = openpyxl.Workbook()
    ws_f = wb_failed.active
    ws_f.title = "Failed Tests"
    ws_f.append(headers)
    for res in results:
        if res["status"] == "Failed":
            ws_f.append([res["test_id"], res["module"], res["test_name"], res["status"], res["execution_time"], res["priority"], res["failure_reason"]])
    wb_failed.save(os.path.join(excel_dir, "Failed_Test_Cases.xlsx"))
    
    wb_passed = openpyxl.Workbook()
    ws_p = wb_passed.active
    ws_p.title = "Passed Tests"
    ws_p.append(headers)
    for res in results:
        if res["status"] == "Passed":
            ws_p.append([res["test_id"], res["module"], res["test_name"], res["status"], res["execution_time"], res["priority"], ""])
    wb_passed.save(os.path.join(excel_dir, "Passed_Test_Cases.xlsx"))
    
    wb_summary = openpyxl.Workbook()
    ws_s = wb_summary.active
    ws_s.title = "Summary Report"
    ws_s.append(["Metric", "Value"])
    ws_s.append(["Total Tests", total])
    ws_s.append(["Passed", passed_count])
    ws_s.append(["Failed", failed_count])
    ws_s.append(["Pass Rate", f"{pass_rate:.2f}%"])
    wb_summary.save(os.path.join(excel_dir, "Summary_Report.xlsx"))
    
    print(f"Generated Excel Reports at {excel_dir}")
